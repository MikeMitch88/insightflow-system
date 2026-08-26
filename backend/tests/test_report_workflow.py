"""Comprehensive test suite for InukaOps Report Generation & Approval Workflow.

Tests:
1. Admin Preview & Validation
2. Admin Confirmation Checklist Enforcement
3. Generation sets PENDING_MANAGER_REVIEW and notifies Manager
4. RBAC Security: Admin/Staff CANNOT approve or reject (HTTP 403 Forbidden)
5. Manager Rejection with mandatory reason sets REVISION_REQUIRED
6. Admin Revision creates Version 2 and resubmits
7. Manager Approval sets APPROVED and notifies Admin
8. 12-sheet Excel export validation
9. KPC Log integration across snapshots
10. Notifications and Audit Trail
"""

import io
import pytest
from fastapi.testclient import TestClient
import openpyxl

from src.main import app
from src.auth.service import create_access_token

client = TestClient(app)

# Helper tokens
ADMIN_TOKEN = create_access_token({"id": 1, "sub": "1", "email": "admin@inukafoundation.org", "name": "Program Administrator", "role": "admin", "status": "active"})
MANAGER_TOKEN = create_access_token({"id": 2, "sub": "2", "email": "grace.w@inukafoundation.org", "name": "Grace Wanjiku", "role": "program_manager", "status": "active"})
OFFICER_TOKEN = create_access_token({"id": 3, "sub": "3", "email": "james.o@inukafoundation.org", "name": "James Otieno", "role": "me_officer", "status": "active"})
LEADERSHIP_TOKEN = create_access_token({"id": 5, "sub": "5", "email": "david.m@inukafoundation.org", "name": "David Mwangi", "role": "leadership", "status": "active"})



def test_admin_preview_and_validation():
    """Admin can preview report and view 4 pillars, KPC log, and validation."""
    resp = client.post("/api/reports/preview", json={
        "reporting_period": "August 2026",
        "reporting_period_id": 7,
        "report_type": "Monthly Donor Report",
        "sections": ["executive_summary", "pillars", "kpc_log", "outcomes", "data_quality", "ai_insights"]
    }, headers={"Authorization": f"Bearer {ADMIN_TOKEN}"})

    assert resp.status_code == 200
    data = resp.json()
    assert data["status"] == "ADMIN_REVIEW"
    assert data["reporting_period"] == "August 2026"
    assert data["data_completeness"] >= 90
    assert data["data_quality"] > 0


    # Verify 4 pillars present
    pillars = data["pillar_performance"]
    assert "scholarship" in pillars or "Scholarship" in pillars.get("pillars", {})
    assert "tech" in pillars or "Tech" in pillars.get("pillars", {})
    assert "vocational" in pillars or "Vocational" in pillars.get("pillars", {})
    assert "plus" in pillars or "Plus" in pillars.get("pillars", {})

    # Verify KPC Log
    assert "kpc_log" in data
    assert len(data["kpc_log"].get("records", [])) > 0

    # Verify Validation
    assert "validation_results" in data
    assert data["validation_results"]["summary"]["passed"] > 0


def test_admin_generation_requires_all_checkboxes():
    """Generation fails with HTTP 400 if not all 4 confirmation checkboxes are ticked."""
    resp = client.post("/api/reports/generate", json={
        "title": "August 2026 Monthly Donor Report",
        "report_type": "Monthly Donor Report",
        "reporting_period": "August 2026",
        "reporting_period_id": 7,
        "confirmed_reviewed": True,
        "confirmed_kpis": True,
        "confirmed_warnings": False,  # Missing one confirmation
        "confirmed_ready": True
    }, headers={"Authorization": f"Bearer {ADMIN_TOKEN}"})

    assert resp.status_code == 400
    assert "confirmation required" in resp.json()["detail"].lower()


def test_full_report_lifecycle_and_rbac():
    """
    Test complete lifecycle:
    1. Admin generates report -> Status: PENDING_MANAGER_REVIEW
    2. Admin attempts to approve/reject -> HTTP 403 Forbidden
    3. Officer attempts to approve/reject -> HTTP 403 Forbidden
    4. Manager rejects with reason -> Status: REVISION_REQUIRED
    5. Admin creates Version 2 -> Status: PENDING_MANAGER_REVIEW
    6. Manager approves -> Status: APPROVED
    """
    # 1. Admin generates report
    gen_resp = client.post("/api/reports/generate", json={
        "title": "August 2026 Monthly Donor Report",
        "report_type": "Monthly Donor Report",
        "reporting_period": "August 2026",
        "reporting_period_id": 7,
        "confirmed_reviewed": True,
        "confirmed_kpis": True,
        "confirmed_warnings": True,
        "confirmed_ready": True
    }, headers={"Authorization": f"Bearer {ADMIN_TOKEN}"})

    assert gen_resp.status_code == 201
    report = gen_resp.json()
    report_id = report["id"]
    assert report["status"] == "PENDING_MANAGER_REVIEW"
    assert report["version"] == 1

    # 2. RBAC: Admin MUST NEVER approve or reject
    admin_approve = client.post(f"/api/reports/{report_id}/approve", json={
        "comment": "Admin trying to approve"
    }, headers={"Authorization": f"Bearer {ADMIN_TOKEN}"})
    assert admin_approve.status_code == 403
    assert "forbidden" in admin_approve.json()["detail"].lower()

    admin_reject = client.post(f"/api/reports/{report_id}/reject", json={
        "reason": "Admin trying to reject"
    }, headers={"Authorization": f"Bearer {ADMIN_TOKEN}"})
    assert admin_reject.status_code == 403
    assert "forbidden" in admin_reject.json()["detail"].lower()

    # 3. RBAC: Other staff (M&E Officer, Leadership) cannot approve
    officer_approve = client.post(f"/api/reports/{report_id}/approve", json={}, headers={"Authorization": f"Bearer {OFFICER_TOKEN}"})
    assert officer_approve.status_code == 403

    leader_reject = client.post(f"/api/reports/{report_id}/reject", json={"reason": "test"}, headers={"Authorization": f"Bearer {LEADERSHIP_TOKEN}"})
    assert leader_reject.status_code == 403

    # 4. Manager rejects with mandatory reason
    bad_reject = client.post(f"/api/reports/{report_id}/reject", json={"reason": ""}, headers={"Authorization": f"Bearer {MANAGER_TOKEN}"})
    assert bad_reject.status_code in [400, 422]


    reject_resp = client.post(f"/api/reports/{report_id}/reject", json={
        "reason": "Vocational completion rate needs verification with Nakuru center."
    }, headers={"Authorization": f"Bearer {MANAGER_TOKEN}"})

    assert reject_resp.status_code == 200
    rejected = reject_resp.json()
    assert rejected["status"] == "REVISION_REQUIRED"
    assert rejected["rejected_by"] == "Grace Wanjiku"
    assert "Vocational completion rate" in rejected["rejection_reason"]

    # 5. Admin creates Version 2
    revise_resp = client.post(f"/api/reports/{report_id}/revise", json={
        "notes": "Verified Nakuru vocational assessment data."
    }, headers={"Authorization": f"Bearer {ADMIN_TOKEN}"})

    assert revise_resp.status_code == 200
    revised = revise_resp.json()
    v2_id = revised["id"]
    assert revised["version"] == 2
    assert revised["status"] == "PENDING_MANAGER_REVIEW"

    # 6. Manager approves Version 2
    approve_resp = client.post(f"/api/reports/{v2_id}/approve", json={
        "comment": "Verified and approved for donor submission."
    }, headers={"Authorization": f"Bearer {MANAGER_TOKEN}"})

    assert approve_resp.status_code == 200
    approved = approve_resp.json()
    assert approved["status"] == "APPROVED"
    assert approved["approved_by"] == "Grace Wanjiku"


def test_12_sheet_excel_download():
    """Verify Excel export contains all 12 requested sheets and structured data."""
    # Generate a report
    gen_resp = client.post("/api/reports/generate", json={
        "title": "August 2026 Donor Impact Report",
        "report_type": "Monthly Donor Report",
        "reporting_period": "August 2026",
        "reporting_period_id": 7,
        "confirmed_reviewed": True,
        "confirmed_kpis": True,
        "confirmed_warnings": True,
        "confirmed_ready": True
    }, headers={"Authorization": f"Bearer {ADMIN_TOKEN}"})
    report_id = gen_resp.json()["id"]

    # Download Excel
    dl_resp = client.get(f"/api/reports/{report_id}/download/excel")
    assert dl_resp.status_code == 200
    assert dl_resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    # Load workbook and verify sheets
    wb = openpyxl.load_workbook(io.BytesIO(dl_resp.content))
    sheet_names = wb.sheetnames

    required_sheets = [
        "Executive Summary",
        "Beneficiaries",
        "Pillar Performance",
        "Scholarship",
        "Plus",
        "Vocational",
        "Tech",
        "Outcomes",
        "Data Quality",
        "KPC Log",
        "AI Insights",
        "Report Metadata"
    ]

    for req_sheet in required_sheets:
        assert req_sheet in sheet_names, f"Missing required sheet: {req_sheet}"

    # Verify KPC Log sheet has entries
    kpc_sheet = wb["KPC Log"]
    assert kpc_sheet.max_row >= 2  # Has header and data rows


def test_notifications_and_audit_timeline():
    """Notifications are created and timeline returns audit records."""
    # Check notifications for admin
    notif_resp = client.get("/api/notifications?role=admin")
    assert notif_resp.status_code == 200
    data = notif_resp.json()
    assert "notifications" in data
    assert len(data["notifications"]) > 0

    # Get a report timeline
    rep_resp = client.get("/api/reports/1")
    assert rep_resp.status_code == 200
    rep_data = rep_resp.json()
    assert "timeline" in rep_data
    assert len(rep_data["timeline"]) > 0
